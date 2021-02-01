import json
import requests
from openpyxl import load_workbook

wb = load_workbook("SampleDCMap.xlsx")
data = wb.active
simplidc_server_url = "http://localhost:8080/api/barzelim"
headers = {'content-type': 'application/json'}


def add_new_racks_to_db(data):
    data_rows = data.max_row
    racks = []

    for row in range(2, data_rows + 1):
        cur_rack = {}
        cur_node = {}
        cur_link = {}
        cur_rack['name'] = data.cell(row, 1).value
        cur_rack['networkId'] = data.cell(row, 22).value
        cur_rack['size'] = 42

        # only if rack does not exist yet create it
        if cur_rack['name'] not in racks:
            racks.append(cur_rack['name'])
            create_racknode(cur_rack)

        # in the same order as the xDevice node
        cur_node['serialNumber'] = "'" + data.cell(row, 6).value + "'"
        cur_node['name'] = data.cell(row, 5).value
        cur_node['vendor'] = data.cell(row, 5).value
        cur_node['osType'] = data.cell(row, 19).value
        cur_node['osVersion'] = data.cell(row, 19).value
        cur_node['rackNumber'] = data.cell(row, 1).value
        cur_node['unumber'] = data.cell(row, 2).value

        if data.cell(row, 4).value == "Storage":
            cur_node['arrayType'] = data.cell(row, 5).value + " " + data.cell(row, 7).value
            # cur_node['arrayType'] = data.cell(row, 7).value
            cur_node['arrayProtocol'] = data.cell(row, 4).value
            cur_node['extramgmtIps'] = data.cell(row, 9).value
            cur_node['clusterName'] = data.cell(row, 5).value

            create_storagenode(cur_node)

        if data.cell(row, 4).value == "Server":
            cur_node['fromFactor'] = ''
            cur_node['externalStorage'] = ''

            create_servernode(cur_node)

        if data.cell(row, 4).value == "Network":
            cur_node['isLayer2'] = False;
            cur_node['isLayer3'] = False;

            create_networknode(cur_node)

        if data.cell(row, 4).value == 'XDevice':
            create_xdevicenode(cur_node)

        put_in_rack(cur_node['serialNumber'], cur_node['rackNumber'])


def create_racknode(rackdata):
    save_to_db_response = requests.post(simplidc_server_url + "/racks", data=json.dumps(rackdata), headers=headers)


def create_storagenode(storagedata):
    save_to_db_response = requests.post(simplidc_server_url + "/devices/storage", data=json.dumps(storagedata),
                                        headers=headers)


def create_networknode(netwrokdata):
    save_to_db_response = requests.post(simplidc_server_url + "/devices/networks", data=json.dumps(netwrokdata),
                                        headers=headers)

def create_servernode(serverdata):
    save_to_db_response = requests.post(simplidc_server_url + "/devices/server", data=json.dumps(serverdata),
                                        headers=headers)


def create_xdevicenode(xdevicedata):
    save_to_db_response = requests.post(simplidc_server_url + "/devices/xdevice", data=json.dumps(xdevicedata),
                                        headers=headers)


def put_in_rack(serialNumber, rackNumber):
    insert_url = simplidc_server_url + "/racks/insert?rackNumber=" + str(rackNumber) + "&serialNumber=" + serialNumber
    save_to_db_response = requests.post(insert_url, headers=headers)


def connect_to_lan(fromSerial, toSerial):
    connect_lan_url = simplidc_server_url + "/lan/connect?fromSerial=" + fromSerial + "&toSerial=" + toSerial
    save_to_db_response = requests.post(connect_lan_url, headers=headers)


def connect_to_san(serverSerial, storageName):
    connect_san_url = simplidc_server_url + "/san/connect?serverSerial=" + serverSerial + "&storageName=" + storageName
    save_to_db_response = requests.post(connect_san_url, headers=headers)


add_new_racks_to_db(data)
